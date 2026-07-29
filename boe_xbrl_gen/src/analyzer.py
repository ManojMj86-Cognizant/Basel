"""
Package Analyzer (Phase 1) — sample-independent enumeration of:
  * templates (tables) and their datapoints, from the DPM annotated-templates workbooks
  * validation rules and the tables each one references, from the validations workbooks
  * the join: per template -> datapoints + relevant rules

This is the shared core consumed by the Streamlit UI and by sample-independent generation.

Inputs (BoE v4.0.0 release folders, defaults to the local extracted locations):
  annotated_dir : ...\\boebankingtaxonomydpmv400        (Annotated Templates *.xlsx)
  validations_dir: ...\\boebankingtaxonomyvalidationsv400 (Validations *.xlsx)
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

ROOT = r"C:\Users\177069\ClaudeLearning"
ANNOTATED_DIR = fr"{ROOT}\boebankingtaxonomydpmv400"
VALIDATIONS_DIR = fr"{ROOT}\boebankingtaxonomyvalidationsv400"

# A template/table code looks like C01.00.01.01, OF08.01.01.01, SR99.01.01.01, LV50.00.00.01
TABLE_CODE_RE = re.compile(r"^[A-Z]{1,3}\d{2,3}(\.\d{2})+$")
TABLE_REF_RE = re.compile(r"\{t:\s*([A-Z]{1,3}\d{2,3}(?:\.\d{2})+)")


# --------------------------------------------------------------------------- rules
@dataclass
class Rule:
    code: str
    framework: str
    tables: set = field(default_factory=set)
    expr: str = ""
    klass: str = ""


def _classify(expr: str) -> str:
    e = expr or ""
    if "matches" in e or "isMatch" in e:
        return "format"
    if re.search(r"\bisNull\b|\bempty\b", e):
        return "existence"
    has_if = bool(re.search(r"\bif\b", e))
    if re.search(r">=|<=|[<>]", e) and "=" not in re.sub(r">=|<=|!=", "", e):
        base = "inequality"
    elif "=" in e:
        if "imax" in e or "imin" in e:
            base = "equality-minmax"
        elif "*" in e or "/" in e:
            base = "equality-scaled"
        elif "+" in e or "sum" in e:
            base = "equality-additivity"
        else:
            base = "equality-simple"
    elif re.search(r">=|<=|[<>]", e):
        base = "inequality"
    else:
        base = "other"
    return ("conditional/" + base) if has_if else base


def parse_validations(validations_dir=VALIDATIONS_DIR) -> list[Rule]:
    rules = []
    for wb_path in Path(validations_dir).glob("*.xlsx"):
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                hdr = list(next(rows))
            except StopIteration:
                continue
            idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
            if "Rule code" not in idx:
                continue
            c_code = idx["Rule code"]
            c_fw = idx.get("Framework version code")
            c_expr = idx.get("Simplified Expression", idx.get("Expression"))
            c_scope = idx.get("Scope")
            for row in rows:
                code = row[c_code] if c_code < len(row) else None
                if not code:
                    continue
                expr = ""
                for c in (c_expr, c_scope):
                    if c is not None and c < len(row) and row[c]:
                        expr += " " + str(row[c])
                tables = set(TABLE_REF_RE.findall(expr))
                rules.append(Rule(
                    code=str(code),
                    framework=str(row[c_fw]) if c_fw is not None and c_fw < len(row) else ws.title,
                    tables=tables, expr=expr.strip(), klass=_classify(expr),
                ))
        wb.close()
    return rules


def rules_by_table(rules):
    out = {}
    for r in rules:
        for t in r.tables:
            out.setdefault(t, []).append(r)
    return out


# ----------------------------------------------------------------------- templates
@dataclass
class Template:
    code: str
    workbook: str
    title: str = ""
    n_rows: int = 0
    n_cols: int = 0


def parse_templates(annotated_dir=ANNOTATED_DIR) -> dict:
    templates = {}
    for wb_path in Path(annotated_dir).glob("*Annotated Templates*.xlsx"):
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            if not TABLE_CODE_RE.match(ws.title):
                continue
            title = ""
            try:
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                title = next((str(c) for c in first if c), "")
            except StopIteration:
                pass
            templates[ws.title] = Template(code=ws.title, workbook=wb_path.name, title=title)
        wb.close()
    return templates


def table_family(code: str) -> str:
    """C01.00.01.01 -> C01.00 (fallback grouping when an exact match isn't found)."""
    parts = code.split(".")
    return ".".join(parts[:2]) if len(parts) >= 2 else code


def template_datapoints(annotated_dir, workbook, sheet, max_rows=400):
    """Extract a template sheet's datapoints: the row codes, column codes, and the
    metric + explicit dimension members annotated on each. Returns a structured dict.

    Annotated-template layout (EBA/BoE): row 3 holds the dimension axis headers
    (col I = 'Metrics', cols J.. = 'eba_dim:XXX'); a data row carries its row code in
    col G, its metric in col I, and dimension members in cols J.. ; column codes appear
    on the header row(s) above the data grid.
    """
    wb = openpyxl.load_workbook(Path(annotated_dir) / workbook, read_only=True, data_only=True)
    ws = wb[sheet]
    grid = [list(r) for r in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True)]
    wb.close()

    def cell(r, c):
        return grid[r][c] if 0 <= r < len(grid) and 0 <= c < len(grid[r]) else None

    # locate the axis header row (the one containing 'Metrics')
    axis_row = None
    for i, row in enumerate(grid):
        if any(isinstance(v, str) and v.strip() == "Metrics" for v in row):
            axis_row = i
            break
    dims = []
    metric_col = None
    if axis_row is not None:
        for j, v in enumerate(grid[axis_row]):
            if not isinstance(v, str):
                continue
            if v.strip() == "Metrics":
                metric_col = j
            m = re.match(r"^([a-z_]+:[A-Z]{2,4})\b", v.strip())
            if m:
                dims.append((j, m.group(1)))

    rows = []
    row_code_col = None
    for j in range(0, 10):           # row code lives in an early column (often G=6)
        codes = [cell(i, j) for i in range(len(grid))]
        if sum(1 for v in codes if isinstance(v, str) and re.fullmatch(r"\d{4}", v.strip())) >= 2:
            row_code_col = j
            break
    if row_code_col is not None and metric_col is not None:
        for i in range(len(grid)):
            rc = cell(i, row_code_col)
            if not (isinstance(rc, str) and re.fullmatch(r"\d{4}", rc.strip())):
                continue
            metric = cell(i, metric_col)
            members = {}
            for (j, dim) in dims:
                v = cell(i, j)
                if isinstance(v, str) and ":" in v:
                    members[dim] = v.strip().split(" ")[0]
            rows.append({"row": rc.strip(),
                         "metric": (metric.strip().split(" ")[0] if isinstance(metric, str) else None),
                         "members": members})
    return {"dimensions": [d for _, d in dims], "rows": rows, "n_rows": len(rows)}


def analyze(annotated_dir=ANNOTATED_DIR, validations_dir=VALIDATIONS_DIR):
    """Top-level API for the UI: returns templates (with relevant rule codes) and rules."""
    rules = parse_validations(validations_dir)
    templates = parse_templates(annotated_dir)
    by_tab = rules_by_table(rules)            # exact table-code -> [Rule]
    result_templates = {}
    for code, t in templates.items():
        relevant = by_tab.get(code, [])
        from collections import Counter
        result_templates[code] = {
            "code": code, "title": t.title, "workbook": t.workbook,
            "rule_codes": [r.code for r in relevant],
            "rule_classes": dict(Counter(r.klass for r in relevant)),
            "n_rules": len(relevant),
        }
    return {"templates": result_templates,
            "rules": {r.code: {"framework": r.framework, "tables": sorted(r.tables),
                               "klass": r.klass, "expr": r.expr} for r in rules}}


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    rules = parse_validations()
    templates = parse_templates()
    by_tab = rules_by_table(rules)
    from collections import Counter
    print(f"rules parsed: {len(rules)}  across frameworks: "
          f"{sorted(set(r.framework for r in rules))}")
    print(f"rule classes: {dict(Counter(r.klass for r in rules))}")
    print(f"templates: {len(templates)}")
    print(f"distinct tables referenced by rules: {len(by_tab)}")
    matched = sum(1 for c in templates if c in by_tab)
    print(f"templates with an exact rule match: {matched}/{len(templates)}")
    print("\nsample templates with EXACT relevant-rule counts:")
    for code in list(templates)[:20]:
        n = len(by_tab.get(code, []))
        print(f"  {code:18} rules={n:4}  {templates[code].title[:48]}")
    # demonstrate datapoint extraction on one template
    demo = next((c for c in templates if c.startswith("C01")), list(templates)[0])
    t = templates[demo]
    dp = template_datapoints(ANNOTATED_DIR, t.workbook, demo)
    print(f"\ndatapoints for {demo}: {dp['n_rows']} rows, dims={dp['dimensions'][:8]}")
    for r in dp["rows"][:5]:
        print(f"  row {r['row']}  metric={r['metric']}  members={r['members']}")
