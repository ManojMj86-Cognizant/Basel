"""Classify the rules in a BoE validations workbook by their Simplified Expression shape,
so we know the solver work distribution for that framework (e.g. banking_reporting / PRA001)."""
import re
import sys
from collections import Counter

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
path = sys.argv[1]
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb[wb.sheetnames[-1]]   # last sheet = the rule table
rows = ws.iter_rows(values_only=True)
hdr = list(next(rows))
idx = {h: i for i, h in enumerate(hdr) if h}
col_expr = idx.get("Simplified Expression", idx.get("Expression"))
col_code = idx.get("Rule code")
col_pre = idx.get("Precondition")

print(f"sheet: {ws.title}  columns: {hdr}")

cls = Counter()
funcs = Counter()
n = 0
examples = {}
for row in rows:
    expr = row[col_expr] if col_expr is not None and col_expr < len(row) else None
    if not expr:
        continue
    n += 1
    e = str(expr)
    has_cond = bool(re.search(r"\bif\b", e))
    for fn in re.findall(r"\b(imax|imin|abs|matches|isMatch|exp|sum|count|empty)\b", e):
        funcs[fn] += 1
    # classify by dominant relation
    if "matches" in e or "isMatch" in e:
        c = "format"
    elif re.search(r"\bempty\b", e):
        c = "existence"
    elif re.search(r">=|<=|[<>]", e) and "=" not in re.sub(r">=|<=", "", e):
        c = "inequality"
    elif ">=" in e or "<=" in e or re.search(r"[<>]", e):
        c = "inequality(+eq?)"
    elif "=" in e:
        if "imax" in e or "imin" in e:
            c = "equality-minmax"
        elif re.search(r"\bexp\(", e):
            c = "equality-tolerance"
        elif "*" in e or "/" in e:
            c = "equality-scaled"
        elif "+" in e or "sum" in e:
            c = "equality-additivity"
        else:
            c = "equality-simple"
    else:
        c = "other"
    if has_cond:
        c = "conditional/" + c
    cls[c] += 1
    if c not in examples:
        examples[c] = (row[col_code], e[:120])

print(f"\ntotal rules with expression: {n}")
print("\nclass distribution:")
for c, k in cls.most_common():
    print(f"  {k:5d}  {c}")
print("\nfunction usage:")
for f, k in funcs.most_common():
    print(f"  {k:5d}  {f}")
print("\nexamples per class:")
for c, (code, ex) in sorted(examples.items()):
    print(f"  [{c}] {code}: {ex}")
