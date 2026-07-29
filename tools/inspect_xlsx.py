import sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
path = sys.argv[1]
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print(f"FILE: {path}")
print(f"SHEETS ({len(wb.sheetnames)}):")
for ws in wb.worksheets:
    try:
        dim = ws.calculate_dimension()
    except Exception:
        dim = "?"
    print(f"  - {ws.title!r}  dims={dim}")

# Dump a preview of each sheet (first ~12 rows, first ~10 cols)
maxprev = int(sys.argv[2]) if len(sys.argv) > 2 else 12
for ws in wb.worksheets:
    print("\n" + "=" * 80)
    print(f"SHEET: {ws.title}")
    print("=" * 80)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=maxprev, values_only=True)):
        cells = ["" if c is None else str(c) for c in row[:10]]
        print(f"r{i+1}: " + " | ".join(cells))
