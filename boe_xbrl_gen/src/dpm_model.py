"""
DPM model loader for the Bank of England Banking v4.0.0 taxonomy.

Parses the DPM dictionary workbook into a reusable model:
  - metrics:    code -> {label, owner, prefix, qname, datatype, period_type,
                         balance, ref_domain_owner, ref_domain_code}
  - dimensions: code -> {label, owner, prefix, qname, domain_owner, domain_code}
  - domains:    code -> {label, type (explicit|typed), datatype, nillable, sheet}
  - members:    "<owner>_<domain>" -> [{code, label, owner, prefix, qname,
                                        usable, default}]

The model is the source of truth for *what values are valid*. It is consumed by the
value engine (assign type-correct random values) and, later, by taxonomy-driven
context generation.

Owner -> namespace-prefix mapping mirrors the instance documents:
  metric:    eba -> eba_met,  boe -> boe_met
  dimension: eba -> eba_dim,  boe -> boe_dim
  member:    eba_BA -> eba_BA, boe_xx -> boe_xx   (prefix == "<owner>_<domain>")
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def _col_index(header_row, *names):
    """0-based index of the first column matching any of `names`, or None if absent.

    Tolerant by design so the loader stays **taxonomy-agnostic**: DPM dictionary workbooks
    differ in which optional columns they carry (e.g. the Insurance Domains sheet has no
    'Is Nillable'). A missing column yields None; `_val(row, None)` then returns None rather
    than crashing. Callers that truly require a column (e.g. 'Code') check for None explicitly.
    """
    norm = {(_clean(h)): i for i, h in enumerate(header_row)}
    for n in names:
        if _clean(n) in norm:
            return norm[_clean(n)]
    return None


def _clean(v):
    return str(v).strip().lower() if v is not None else ""


def _val(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _truthy(v):
    return str(v).strip().lower() in ("true", "1", "yes", "x")


def load_dpm(dict_path: str) -> dict:
    wb = openpyxl.load_workbook(dict_path, read_only=True, data_only=True)

    model = {"metrics": {}, "dimensions": {}, "domains": {}, "members": {}}

    # ---- Metrics ----
    ws = wb["Metrics"]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    c_code = _col_index(hdr, "Code")
    c_label = _col_index(hdr, "Label (en)", "Label")
    c_owner = _col_index(hdr, "Owner")
    c_dtype = _col_index(hdr, "Data type")
    c_period = _col_index(hdr, "Period type")
    c_balance = _col_index(hdr, "Balance type")
    c_rdo = _col_index(hdr, "Referenced domain owner")
    c_rdc = _col_index(hdr, "Referenced domain code")
    seen = set()
    for row in rows:
        code = _val(row, c_code)
        if not code or code in seen:
            continue
        seen.add(code)
        owner = (_val(row, c_owner) or "eba").lower()
        prefix = f"{owner}_met"
        model["metrics"][code] = {
            "label": _val(row, c_label),
            "owner": owner,
            "prefix": prefix,
            "qname": f"{prefix}:{code}",
            "datatype": (_val(row, c_dtype) or "").upper(),
            "period_type": _val(row, c_period),
            "balance": _val(row, c_balance),
            "ref_domain_owner": _val(row, c_rdo),
            "ref_domain_code": _val(row, c_rdc),
        }

    # ---- Dimensions ----
    ws = wb["Dimensions"]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    c_code = _col_index(hdr, "Code")
    c_label = _col_index(hdr, "Label (en)", "Label")
    c_owner = _col_index(hdr, "Owner")
    c_do = _col_index(hdr, "Domain owner")
    c_dc = _col_index(hdr, "Domain code")
    for row in rows:
        code = _val(row, c_code)
        if not code:
            continue
        owner = (_val(row, c_owner) or "eba").lower()
        prefix = f"{owner}_dim"
        model["dimensions"][code] = {
            "label": _val(row, c_label),
            "owner": owner,
            "prefix": prefix,
            "qname": f"{prefix}:{code}",
            "domain_owner": (_val(row, c_do) or "").lower() or None,
            "domain_code": _val(row, c_dc),
        }

    # ---- Domains ----
    ws = wb["Domains"]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    c_code = _col_index(hdr, "Code")
    c_label = _col_index(hdr, "Label (en)", "Label")
    c_type = _col_index(hdr, "Type")
    c_owner = _col_index(hdr, "Owner")
    c_dtype = _col_index(hdr, "Data type")
    c_nil = _col_index(hdr, "Is Nillable")
    for row in rows:
        code = _val(row, c_code)
        if not code:
            continue
        owner = (_val(row, c_owner) or "eba").lower()
        model["domains"][code] = {
            "label": _val(row, c_label),
            "owner": owner,
            "type": (_val(row, c_type) or "").lower(),   # explicit | typed
            "datatype": (_val(row, c_dtype) or "").upper() or None,
            "nillable": _truthy(_val(row, c_nil)),
            "sheet": f"{owner}_{code}",
        }

    # ---- Members (one sheet per explicit domain, named "<owner>_<domain>") ----
    known_meta = {"Note", "Owners", "Domains", "Dimensions", "Metrics"}
    for sheet in wb.sheetnames:
        if sheet in known_meta:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            hdr = next(rows)
        except StopIteration:
            continue
        c_code = _col_index(hdr, "Code")
        if c_code is None:                 # not a member sheet (no Code column)
            continue
        c_label = _col_index(hdr, "Label (en)", "Label")
        c_owner = _col_index(hdr, "Owner")
        c_default = _col_index(hdr, "Default")
        c_usable = _col_index(hdr, "Usable")
        members = []
        seen_m = set()
        for row in rows:
            code = _val(row, c_code)
            if not code or code in seen_m:
                continue
            seen_m.add(code)
            owner = (_val(row, c_owner) or sheet.split("_")[0]).lower()
            members.append({
                "code": code,
                "label": _val(row, c_label),
                "owner": owner,
                "prefix": sheet,
                "qname": f"{sheet}:{code}",
                "usable": _truthy(_val(row, c_usable)) if c_usable is not None else True,
                "default": _truthy(_val(row, c_default)) if c_default is not None else False,
            })
        if members:
            model["members"][sheet] = members

    return model


def main():
    dict_path = sys.argv[1] if len(sys.argv) > 1 else (
        r"C:\Users\177069\ClaudeLearning\boebankingtaxonomydpmv400"
        r"\Bank of England Banking DPM dictionary v4.0.0.xlsx"
    )
    out = sys.argv[2] if len(sys.argv) > 2 else (
        r"C:\Users\177069\ClaudeLearning\boe_xbrl_gen\model\dpm_model.json"
    )
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    model = load_dpm(dict_path)

    Path(out).parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(model, fh, ensure_ascii=False, indent=1)

    # ---- summary ----
    from collections import Counter
    dtypes = Counter(m["datatype"] for m in model["metrics"].values())
    dom_types = Counter(d["type"] for d in model["domains"].values())
    n_members = sum(len(v) for v in model["members"].values())
    print(f"metrics:     {len(model['metrics'])}")
    print(f"dimensions:  {len(model['dimensions'])}")
    print(f"domains:     {len(model['domains'])}  ({dict(dom_types)})")
    print(f"member sheets: {len(model['members'])}, total members: {n_members}")
    print(f"\nmetric datatypes: {dict(dtypes)}")
    typed = {c: d for c, d in model["domains"].items() if d["type"] == "typed"}
    print(f"\ntyped domains ({len(typed)}): " +
          ", ".join(f"{c}:{d['datatype']}" for c, d in list(typed.items())[:30]))
    print(f"\nsaved model -> {out}")


if __name__ == "__main__":
    main()
