"""Rule-driven generation engine — Stage 1: parse the BoE **validations workbook** rules and
resolve their cell refs to real datapoints via the rc-code bridge.

The workbook expresses each rule over **cell refs** in BoE r/c/z codes, e.g.
    isum({t: C14.01.01.01, c: 0430, z: 0010; 0020; 0030; 0040, seq: True, id: v1})
      i= {t: C13.01.01.01, r: 0010, c: 0890, id: v2}
       i+ {t: OF19.00.01.01, r: 0010, c: 0570, id: v3} i* 12.5
       i+ {t: OF20.00.01.01, r: 0010, c: 0410, id: v4} i* 12.5 ...

We parse the **additive-equality** shape (the dominant rule type: a side `i=` a sum/difference
of cell terms, each optionally `i* const`, optionally wrapped in `isum(...)`), and resolve each
cell ref to its `(metric, dims)` datapoint(s) using `table_model.rc_codes` + the table linkbase
positions. This sidesteps `resolver.bind` (which doesn't match the studio's built facts) — we
place/compute values at the *exact* cells a rule references.

Rules we can't parse into the additive-equality shape (inequalities, conditionals, imax/iif,
preconditions, etc.) are returned with `parsed=None` so the caller can skip or defer them.
"""
from __future__ import annotations

import os
import re

import table_model

# Max cells in a connected component before `solve_cells_lp` skips its dense per-cell ≥0 LP (which
# scales as a (~2·nf+|cells|)×(2·nf) dense matrix). Above this, exact equalities are still derived
# via RREF but free vars get random non-negative values. Tunable via env for experimentation.
# Default 20000 covers PRA001's biggest components (OF07 ~17.8k cells, OF08 ~11.7k) so the ≥0 LP
# runs on them — verified 2026-06-26: cap=2500 left 2,362 negatives; cap=20000 → 144 (the hard tail),
# 0 dim-invalid, in ~177 s. A lower cap is faster but ships a non-shippable file (fails the ≥0 rules).
_LP_CELL_CAP = int(os.environ.get("GENVALID_LP_CELL_CAP", "20000"))

# Set GENVALID_NO_LP=1 to skip the scipy LP entirely (exact RREF additive solve + random non-negative
# free vars; inequality rules best-effort). Also auto-engaged when scipy can't be imported.
_NO_LP_ENV = os.environ.get("GENVALID_NO_LP", "0").strip().lower() not in ("", "0", "false", "no")

_LINPROG = "unset"   # cache: "unset" | callable | None


def _get_linprog():
    """Import scipy.optimize.linprog behind a watchdog thread + timeout. Returns the function, or
    None if scipy is unavailable / its import hangs (a known environment failure mode on this box —
    scipy's import can deadlock during BLAS init). Cached after the first attempt."""
    global _LINPROG
    if _LINPROG != "unset":
        return _LINPROG
    if _NO_LP_ENV:
        _LINPROG = None
        return None
    import threading
    box: dict = {}

    def _imp():
        try:
            from scipy.optimize import linprog
            box["fn"] = linprog
        except Exception as e:                       # pragma: no cover
            box["err"] = e
    th = threading.Thread(target=_imp, daemon=True)
    th.start()
    th.join(float(os.environ.get("GENVALID_SCIPY_IMPORT_TIMEOUT", "25")))
    _LINPROG = box.get("fn")                          # None if the import hung past the timeout
    return _LINPROG

# ----------------------------------------------------------------- workbook loading
def load_workbook_rules(xlsx_path: str, sheet: str | None = None) -> list[dict]:
    """Return [{code, expression, simplified, precondition, include, deactivated, severity,
    tables}] for every rule row of the framework sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[-1]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c) if c is not None else "" for c in rows[0]]

    def idx(name, default=None):
        return hdr.index(name) if name in hdr else default

    ci = idx("Rule code"); ce = idx("Expression"); cs = idx("Simplified Expression")
    cp = idx("Precondition"); cinc = idx("Include in XBRL"); cdea = idx("Deactivated")
    csev = idx("Severity and modules"); csc = idx("Scope"); cwh = idx("Where"); cjn = idx("Join")
    tcols = [idx(f"T{i}") for i in range(1, 6) if idx(f"T{i}") is not None]

    def cell(r, i):
        return str(r[i]) if i is not None and i < len(r) and r[i] else ""

    out = []
    for r in rows[1:]:
        code = r[ci] if ci is not None and ci < len(r) else None
        if not code:
            continue
        tabs = [str(r[t]) for t in tcols if t < len(r) and r[t]]
        out.append({
            "code": str(code).strip(),
            "expression": cell(r, ce),
            "simplified": cell(r, cs),
            "precondition": cell(r, cp),
            "scope": cell(r, csc),        # (rows × z) the column expression iterates over
            "where": cell(r, cwh),
            "join": cell(r, cjn),
            "include": (str(r[cinc]).strip().lower() == "yes") if cinc is not None and r[cinc] else False,
            "deactivated": (str(r[cdea]).strip().lower() == "yes") if cdea is not None and r[cdea] else False,
            "severity": cell(r, csev),
            "tables": tabs,
        })
    return out


# ----------------------------------------------------------------- expression parsing
_CELL_RE = re.compile(r"\{([^}]*)\}")
_REL_RE = re.compile(r"\bi(=|>=|<=|>|<|!=)\b|\si(=|>=|<=|>|<|!=)\s")


def _parse_cell(body: str) -> dict:
    """'t: C14.01.01.01, c: 0430, z: 0010; 0020, id: v1, seq: True' -> structured cell ref."""
    d = {}
    for part in body.split(","):
        if ":" not in part:
            continue
        k, v = part.split(":", 1)
        d[k.strip()] = v.strip()
    z = d.get("z", "")
    zlist = [x.strip() for x in z.split(";") if x.strip()] if z else []
    return {
        "table": d.get("t", ""),
        "r": d.get("r") or None,
        "c": d.get("c") or None,
        "z": zlist,
        "id": d.get("id"),
        "seq": str(d.get("seq", "")).lower() == "true",
        "dv": d.get("dv"),
    }


def parse_expression(expr: str) -> dict | None:
    """Parse the additive-equality shape into {op, lhs:[term], rhs:[term]} where each term is
    {cell, coef, sum}. Returns None if the expression isn't a plain additive (in)equality.

    A term is `[isum(]{cellref}[)] [i* k]` joined by `i+` / `i-` (top level)."""
    if not expr or "{" not in expr:
        return None
    # Reject NON-additive expressions — otherwise the additive parser silently mis-reads them and
    # emits garbage equations (e.g. b0360 'r0020 = exp(Σcell²,1,2)' parsed as 'r0030 = r0030';
    # b0676 'if c0020≠0 then c0040 = c0030·imax(c0010/c0020,1)' parsed as harmful 'c0020 = c0030').
    # Pure additive uses only {cell}, isum(, i+, i-, i= / i<= / i>=, i* <number>.
    el = expr.lower()
    if any(tok in el for tok in ("exp(", "imax", "imin", "if ", "if(", " then ", "else")):
        return None
    if re.search(r"\}\s*\*\s*\{", expr):      # cell × cell (non-linear) — but number × cell is a coef, OK
        return None
    m = _REL_RE.search(expr)
    if not m:
        return None
    op = "i" + (m.group(1) or m.group(2))
    lhs_s, rhs_s = expr[:m.start()], expr[m.end():]

    def parse_side(s: str):
        # split into additive terms at top-level i+/i- (cell bodies have no i+/i-)
        s = s.strip()
        # normalise leading sign
        tokens = re.split(r"\s+i([+\-])\s+", s)
        # tokens: [term0, sign1, term1, sign2, term2, ...]
        terms = []
        sign = 1.0
        first = tokens[0]
        rest = tokens[1:]
        chunks = [(1.0, first)]
        for i in range(0, len(rest), 2):
            sg = -1.0 if rest[i] == "-" else 1.0
            chunks.append((sg, rest[i + 1]))
        for sg, chunk in chunks:
            cm = _CELL_RE.search(chunk)
            if not cm:
                return None
            cell = _parse_cell(cm.group(1))
            cell["sum"] = "isum" in chunk[:cm.start()].lower()
            # numeric coefficient. Two forms, in EITHER order relative to the cell/function:
            #   * LEADING  'k i* …' / 'k * …' / 'k / …'  (e.g. b0471 '0.9 * {cell}', b0826 '1.2 i* isum(')
            #   * TRAILING '… i* k' / '… * k' / '… / k'  (e.g. v7380 '{cell} i* 12.5')
            # Remove the {..} cell body first so its r/c coords (0020, …) aren't read as coefficients.
            coef = sg
            outside = _CELL_RE.sub(" ", chunk)
            # allow a leading '(' before the coefficient — rules write '(0.9 * {cell})' with the
            # multiplier PARENTHESISED (b0471: c0150 i- (0.9 * c0160) i- (0.8 * c0170) …). Without
            # skipping '(', the 0.9/0.8/… magnitudes were dropped (coef stayed ±1) → wrong equation.
            mlead = re.match(r"[\s(]*-?\s*([0-9]*\.?[0-9]+)\s*i?([*/])", outside)
            if mlead:
                k = float(mlead.group(1))
                coef = coef * k if mlead.group(2) == "*" else coef / k
                outside = outside[mlead.end():]          # consume so the trailing scan won't re-count it
            for km in re.finditer(r"i?([*/])\s*([0-9]*\.?[0-9]+)", outside):
                k = float(km.group(2))
                coef = coef * k if km.group(1) == "*" else coef / k
            terms.append({"cell": cell, "coef": coef})
        return terms

    lhs, rhs = parse_side(lhs_s), parse_side(rhs_s)
    if lhs is None or rhs is None:
        return None
    return {"op": op, "lhs": lhs, "rhs": rhs}


def _semi(v) -> list:
    return [x.strip() for x in str(v).split(";") if x.strip()] if v else []


def parse_scope(scope: str) -> dict | None:
    """`scope({t: OF08.01.01.01, r:0001;.., c:0020;.., z:0001;..})` -> {table, rows, cols, z}.
    Also handles the MULTI-TABLE form `scope((t: A, z:..),(t: B, z:..))` (cross-table rules) by
    unioning the rows/cols/z across clauses — the expression cells carry their own table/r/c, so the
    scope mainly supplies the shared z. Returns None when there is no scope clause."""
    if not scope:
        return None
    if "{" in scope and "}" in scope:
        bodies = [scope[scope.index("{") + 1: scope.rindex("}")]]
    else:
        bodies = re.findall(r"\(([^()]*)\)", scope)        # each (t:.., z:..) clause
    if not bodies:
        return None
    table = ""; rows: set = set(); cols: set = set(); zs: set = set()
    for body in bodies:
        d = {}
        for part in body.split(","):
            if ":" in part:
                k, v = part.split(":", 1)
                d[k.strip()] = v.strip()
        table = table or d.get("t", "")
        rows |= set(_semi(d.get("r", ""))); cols |= set(_semi(d.get("c", ""))); zs |= set(_semi(d.get("z", "")))
    return {"table": table, "rows": sorted(rows), "cols": sorted(cols), "z": sorted(zs)}


def expand_scoped_asts(rule: dict) -> list:
    """Expand one scoped additive rule into a concrete AST per (scope-row × scope-col × scope-z),
    with every term's cell fully specified (single r, c, z) so it resolves to exactly one built
    fact. The expression pins one axis (e.g. columns for `c0090 = c0020+..`, or rows for
    `r0010 = r0070+..`); the scope enumerates the others. Multi-cell terms (`c: 0020; 0070; 0080`)
    are split into one term per value. Returns [] for non-additive / unscoped rules."""
    ast = parse_expression(rule.get("expression", ""))
    if not ast:
        return []
    # Scoped rules enumerate rows/cols/z separately; INLINE rules write r AND c straight into each
    # cell (no Scope clause) — those expand to a single equation from the cells' own r/c/z.
    sc = parse_scope(rule.get("scope", "")) or {"table": "", "rows": [], "cols": [], "z": []}
    table = sc["table"]
    out = []
    for srow in (sc["rows"] or [None]):
        for scol in (sc["cols"] or [None]):
            for sz in (sc["z"] or [None]):
                concrete = {"op": ast["op"], "lhs": [], "rhs": []}
                for side in ("lhs", "rhs"):
                    for term in ast[side]:
                        tc = term["cell"]
                        rvals = _semi(tc.get("r")) or [tc.get("r") or srow]   # term axis, else scope axis
                        cvals = _semi(tc.get("c")) or [tc.get("c") or scol]
                        tz = tc.get("z") or ([sz] if sz else [])
                        for rv in rvals:
                            for cv in cvals:
                                concrete[side].append({
                                    "cell": {"table": tc.get("table") or table,
                                             "r": rv, "c": cv, "z": list(tz)},
                                    "coef": term["coef"]})
                out.append(concrete)
    return out


def isnull_cells(rule: dict) -> list:
    """For an `isNull(...)` rule (a cell that must be EMPTY for the scoped/inline rows·cols·z, e.g.
    b1039 'columns 0101/0102/0103 are null for non-slotting exposure classes'), return the concrete
    `{table, r, c, z}` cells that must NOT be populated — the condition is encoded by those rows, so
    emitting no fact there satisfies the rule. Returns [] for non-isNull rules."""
    expr = rule.get("expression", "")
    if not expr.strip().startswith("isNull"):
        return []
    sc = parse_scope(rule.get("scope", ""))
    out = []
    for m in _CELL_RE.finditer(expr):
        c = _parse_cell(m.group(1))
        rows = _semi(c.get("r")) or (sc["rows"] if sc else []) or [None]
        cols = _semi(c.get("c")) or (sc["cols"] if sc else []) or [None]
        zs = c.get("z") or (sc["z"] if sc else []) or [None]
        for r in rows:
            for col in cols:
                for z in zs:
                    out.append({"table": c.get("table") or (sc["table"] if sc else ""),
                                "r": r, "c": col, "z": [z] if z else []})
    return out


# ----------------------------------------------------------------- cell -> datapoint resolver
class CellResolver:
    """Resolves (table, r/c/z code) -> [(concept, dims)] datapoints via the rc-code bridge.
    Built once per package; caches each table's rc-code index."""

    def __init__(self, extracted_dir: str):
        self.dir = extracted_dir
        self._rend = {t["code"].upper(): t["path"] for t in table_model.list_tables(extracted_dir)}
        self._cache: dict[str, dict] = {}

    def _index(self, table: str) -> dict:
        key = table.upper()
        if key in self._cache:
            return self._cache[key]
        rend = self._rend.get(key)
        byaxis: dict = {"x": {}, "y": {}, "z": {}}
        if rend and os.path.exists(rend):
            p = table_model.parse_table(rend)
            rc = table_model.rc_codes(rend)
            for ax, poss in p.get("axis_positions", {}).items():
                d = {}
                for pos in poss:
                    code = rc.get(pos.get("node"))
                    if code:
                        d[code] = pos
                byaxis[ax] = d
        self._cache[key] = byaxis
        return byaxis

    def resolve(self, cell: dict) -> list[dict]:
        """cell = {table, r, c, z:[...]}. Returns the list of {concept, dims} datapoints it
        denotes (a sequence when z has multiple values or seq=True)."""
        bx = self._index(cell["table"])
        zvals = cell["z"] or [None]
        out = []
        for z in zvals:
            parts = []
            if cell.get("c") is not None:
                parts.append(bx["x"].get(cell["c"]))
            if cell.get("r") is not None:
                parts.append(bx["y"].get(cell["r"]))
            if z is not None:
                parts.append(bx["z"].get(z))
            if any(p is None for p in parts) or not parts:
                continue
            concept, dims = None, {}
            for pos in parts:
                concept = pos.get("concept") or concept
                dims = {**dims, **pos.get("dims", {})}
            if concept:
                out.append({"concept": concept, "dims": dims, "table": cell["table"]})
        return out


# ----------------------------------------------------------------- cell-space solver (Stage 2)
def _dpkey(dp: dict) -> tuple:
    return (dp["concept"], tuple(sorted(dp["dims"].items())))


def plan_equality(ast: dict, resolver: "CellResolver") -> dict | None:
    """Turn an `i=` rule into the aggregated linear equation `Σ coef_k · v_k = 0` (lhs +, rhs -).
    Returns {agg:{key->coef}, preferred:[keys on a single-term 'total' side], dp:{key->dp}} so the
    solver can pick ANY free cell as the derived target — not just the first claim. None if a cell
    didn't resolve."""
    if ast["op"] != "i=":
        return None
    agg: dict = {}               # key -> summed signed coefficient
    dp_by_key: dict = {}
    side_keys: dict = {"lhs": [], "rhs": []}
    for side, sign in (("lhs", 1.0), ("rhs", -1.0)):
        for term in ast[side]:
            dps = resolver.resolve(term["cell"])
            if not dps:
                return None      # a cell didn't resolve -> skip this rule
            for dp in dps:
                k = _dpkey(dp)
                agg[k] = agg.get(k, 0.0) + sign * term["coef"]
                dp_by_key[k] = dp
                side_keys[side].append(k)
    agg = {k: c for k, c in agg.items() if abs(c) > 1e-9}
    if not agg:
        return None
    preferred = [k for side in ("lhs", "rhs") if len(ast[side]) == 1
                 for k in side_keys[side] if k in agg]    # a lone 'total' cell is the natural target
    return {"agg": agg, "preferred": preferred, "dp": dp_by_key}


def solve_cells_linear(asts: list, resolver: "CellResolver", rng, datatype_of=None,
                       null_keys: set | None = None, fixed_values: dict | None = None,
                       present_keys: set | None = None) -> dict:
    """EXACT simultaneous solve of the additive equality system via sparse Gaussian elimination
    (RREF over Fractions). Unlike the greedy `solve_cells`, this satisfies EVERY consistent equation
    at once — needed for over-determined 2D tables (row-sums AND column-sums, multiple decompositions
    of the same total, e.g. OF08 b0744/b0745/b0746). Free variables get datatype-valid randoms;
    dependent variables are derived exactly. Constants (held out of the variable set, moved to RHS):
    `null_keys` (=0), `fixed_values` (pinned), and — if `present_keys` is given — any referenced cell
    NOT generated (absent ⇒ 0, matching the instance). Returns {dp_key -> {concept,dims,value}}."""
    from fractions import Fraction as F
    null_keys = null_keys or set()
    fixed_values = fixed_values or {}

    def is_const(k):
        return k in null_keys or k in fixed_values or (present_keys is not None and k not in present_keys)

    def const_val(k):
        return F(fixed_values[k]).limit_denominator(1) if k in fixed_values else F(0)

    dp_by_key: dict = {}
    eqs: list = []                       # each: (coefs {var:Fraction}, b Fraction)  meaning Σ coef·var = b
    for a in asts:
        if a["op"] != "i=":
            continue
        row: dict = {}
        for side, sgn in (("lhs", 1), ("rhs", -1)):
            for term in a[side]:
                for dp in resolver.resolve(term["cell"]):
                    k = _dpkey(dp); dp_by_key[k] = dp
                    row[k] = row.get(k, F(0)) + F(sgn) * F(term["coef"]).limit_denominator(1000000)
        coefs: dict = {}; b = F(0)
        for k, c in row.items():
            if c == 0:
                continue
            if is_const(k):
                b -= c * const_val(k)        # move constant term to the RHS
            else:
                coefs[k] = coefs.get(k, F(0)) + c
        coefs = {k: c for k, c in coefs.items() if c != 0}
        if coefs:
            eqs.append((coefs, b))

    pivots: dict = {}                    # pivot var -> (coefs incl. pivot, b)  fully reduced
    for coefs, b in eqs:
        coefs = dict(coefs)
        for v in list(coefs):            # reduce by existing pivots
            if v in pivots and coefs.get(v, 0) != 0:
                f = coefs[v]; pc, pb = pivots[v]
                for k, c in pc.items():
                    coefs[k] = coefs.get(k, F(0)) - f * c
                b -= f * pb
                coefs = {k: c for k, c in coefs.items() if c != 0}
        if not coefs:
            continue                     # 0 = b ; consistent (b≈0) or unsatisfiable — skip
        pv = min(coefs); piv = coefs[pv]
        nc = {k: c / piv for k, c in coefs.items()}; nb = b / piv
        for p2, (c2, b2) in list(pivots.items()):   # back-reduce existing pivots by the new one
            if pv in c2 and c2[pv] != 0:
                f = c2[pv]
                merged = {k: c2.get(k, F(0)) - f * c for k, c in nc.items()}
                for k in c2:
                    merged[k] = merged.get(k, c2[k]) if k in nc else c2[k]
                pivots[p2] = ({k: c for k, c in merged.items() if c != 0}, b2 - f * nb)
        pivots[pv] = (nc, nb)

    all_vars = set()
    for coefs, _ in eqs:
        all_vars |= set(coefs)

    def rand_val(k):
        dt = (datatype_of(dp_by_key[k]["concept"].split(":")[-1]) if datatype_of else None) or "MONETARY"
        if dt in ("MONETARY",):
            return F(rng.randint(1, 9999) * 1000)
        if dt == "INTEGER":
            return F(rng.randint(0, 100000))
        if dt == "PERCENTAGE":
            return F(rng.randint(0, 10000), 10000)
        return F(rng.randint(0, 999900), 100)

    val: dict = {}
    for k in all_vars:
        if k not in pivots:
            val[k] = rand_val(k)
    for pv, (nc, nb) in pivots.items():          # pivot row reduced to pivot + free vars only
        val[pv] = nb - sum(c * val.get(k, F(0)) for k, c in nc.items() if k != pv)

    out: dict = {}
    for k, dp in dp_by_key.items():
        if not is_const(k) and k in val:
            out[k] = {"concept": dp["concept"], "dims": dp["dims"], "value": float(val[k])}
    return out


def solve_cells_lp(asts: list, resolver: "CellResolver", rng, datatype_of=None,
                   null_keys: set | None = None, fixed_values: dict | None = None,
                   present_keys: set | None = None, le_constraints: list | None = None) -> dict:
    """Like `solve_cells_linear` (exact additive solve per connected component) but ALSO satisfies
    LINEAR INEQUALITIES. `le_constraints` is a list of `(coef_dict {key->coef}, rhs)` each meaning
    `Σ coef·value(cell) ≤ rhs` — this models A≤B, A≤0, A≥B+C, |A|≤|B| (with cells ≥0), and folds an
    absent/const cell into `rhs` (e.g. b0960 r0343 ≤ r0341 where r0341 is absent → r0343 ≤ 0). Every
    cell is forced ≥0 too. Components touched by a constraint are solved with an LP (scipy HiGHS,
    L1-close to a random target) keeping equalities EXACT; others use random frees."""
    import numpy as np
    from fractions import Fraction as F
    from collections import defaultdict
    linprog = _get_linprog()                          # None if scipy unavailable/hung → no-LP fallback
    null_keys = null_keys or set(); fixed_values = fixed_values or {}; le_constraints = le_constraints or []

    def is_const(k):
        return k in null_keys or k in fixed_values or (present_keys is not None and k not in present_keys)

    def const_val(k):
        return float(fixed_values.get(k, 0))

    dp_by_key: dict = {}
    eqs: list = []                                   # (coefs {var:Fraction}, b Fraction)
    for a in asts:
        if a["op"] != "i=":
            continue
        row: dict = {}
        for side, sgn in (("lhs", 1), ("rhs", -1)):
            for term in a[side]:
                for dp in resolver.resolve(term["cell"]):
                    k = _dpkey(dp); dp_by_key[k] = dp
                    row[k] = row.get(k, F(0)) + F(sgn) * F(term["coef"]).limit_denominator(10**6)
        coefs: dict = {}; b = F(0)
        for k, c in row.items():
            if c == 0:
                continue
            if is_const(k):
                b -= c * F(const_val(k)).limit_denominator(1)
            else:
                coefs[k] = coefs.get(k, F(0)) + c
        coefs = {k: c for k, c in coefs.items() if c != 0}
        if coefs:
            eqs.append((coefs, b))

    # normalise each le constraint: fold const/absent cells into rhs, keep present-cell coefs
    le: list = []                                    # (coef_dict over non-const cells, rhs)
    for cd, rhs in le_constraints:
        cc = {}; r = float(rhs)
        for k, co in cd.items():
            if is_const(k):
                r -= co * const_val(k)
            else:
                cc[k] = cc.get(k, 0.0) + co
        cc = {k: co for k, co in cc.items() if abs(co) > 1e-12}
        if cc:
            le.append((cc, r))

    parent: dict = {}
    def find(x):
        parent.setdefault(x, x); r = x
        while parent[r] != r:
            r = parent[r]
        while parent[x] != r:
            parent[x], x = r, parent[x]
        return r
    def union(a, b):
        parent[find(a)] = find(b)
    for coefs, _ in eqs:
        ks = list(coefs)
        for k in ks[1:]:
            union(ks[0], k)
    for cd, _ in le:
        ks = list(cd)
        for k in ks[1:]:
            union(ks[0], k)

    comp_eqs = defaultdict(list); comp_cells = defaultdict(set); comp_le = defaultdict(list)
    for coefs, b in eqs:
        c = find(next(iter(coefs))); comp_eqs[c].append((coefs, b)); comp_cells[c] |= set(coefs)
    for cd, r in le:
        c = find(next(iter(cd))); comp_le[c].append((cd, r)); comp_cells[c] |= set(cd)

    def rand_val(k):
        dt = (datatype_of(k[0].split(":")[-1]) if datatype_of else None) or "MONETARY"   # k[0]=concept qname
        if dt in ("MONETARY",):
            return F(rng.randint(1, 9999) * 1000)
        if dt == "INTEGER":
            return F(rng.randint(0, 100000))
        if dt == "PERCENTAGE":
            return F(rng.randint(0, 10000), 10000)
        return F(rng.randint(0, 999900), 100)

    def rref(rows):
        piv: dict = {}
        for coefs, b in rows:
            coefs = dict(coefs)
            for v in list(coefs):
                if v in piv and coefs.get(v, 0) != 0:
                    f = coefs[v]; pc, pb = piv[v]
                    for k, c in pc.items():
                        coefs[k] = coefs.get(k, F(0)) - f * c
                    b -= f * pb
                    coefs = {k: c for k, c in coefs.items() if c != 0}
            if not coefs:
                continue
            pv = min(coefs); pcf = coefs[pv]
            nc = {k: c / pcf for k, c in coefs.items()}; nb = b / pcf
            for p2, (c2, b2) in list(piv.items()):
                if pv in c2 and c2[pv] != 0:
                    f = c2[pv]; m = {k: c2.get(k, F(0)) - f * c for k, c in nc.items()}
                    for k in c2:
                        m[k] = m.get(k, c2[k]) if k in nc else c2[k]
                    piv[p2] = ({k: c for k, c in m.items() if c != 0}, b2 - f * nb)
            piv[pv] = (nc, nb)
        return piv

    val: dict = {}
    for comp in set(comp_eqs) | set(comp_le):       # include le-only components (no additive eqs)
        piv = rref(comp_eqs.get(comp, []))
        cells = comp_cells[comp]
        free = [k for k in cells if k not in piv]
        if not free:                                # fully determined — just derive
            for pv, (nc, nb) in piv.items():
                val[pv] = nb - sum(c * val.get(k, F(0)) for k, c in nc.items() if k != pv)
            continue
        # SIZE CAP: the per-cell blanket-≥0 LP builds a dense (~2·nf+|cells|)×(2·nf) matrix, so a
        # huge component (e.g. the 7k-cell OF07/OF08 2D-additive block that the 'total' cells fuse
        # together) makes scipy choke — minutes-to-hang. Above the cap, keep the EXACT equalities
        # (RREF derivation) but skip the LP: assign random NON-NEGATIVE free vars (rand_val ≥ 0) and
        # derive pivots. Equalities still hold exactly; only the ≤/≥ inequalities for this big block
        # are best-effort (that block was the known over-determined hard tail anyway).
        if linprog is None or len(cells) > _LP_CELL_CAP:
            for k in free:
                val[k] = rand_val(k)
            for pv, (nc, nb) in piv.items():
                val[pv] = nb - sum(c * val.get(k, F(0)) for k, c in nc.items() if k != pv)
            continue
        # LP over the free vars for EVERY component: value(cell)=offset+coef·f, every cell ≥ 0 (so the
        # many 'cell ≥ 0' rules hold and |·| in abs-inequalities drops), plus any A≤B for this
        # component; objective = L1-close to a random target so values stay varied. Equalities stay
        # exact (derived from f). Falls back to random frees if the LP is infeasible.
        fidx = {k: i for i, k in enumerate(free)}; nf = len(free)
        def vec(k):
            v = np.zeros(nf); off = 0.0
            if k in fidx:
                v[fidx[k]] = 1.0
            elif k in piv:
                nc, nb = piv[k]; off = float(nb)
                for j, co in nc.items():
                    if j != k and j in fidx:
                        v[fidx[j]] -= float(co)
            return v, off
        # Constraint rows over the free vars f: (coef-array, rhs) meaning coef·f ≤ rhs.
        # Blanket cell ≥ 0 (exposures are non-negative + many 'cell ≥ 0' rules), and every le.
        crows = []
        for k in cells:                              # cell ≥ 0  →  -coef·f ≤ off
            v, off = vec(k); crows.append((-v, off))
        for cd, rhs in comp_le[comp]:                # Σ coef·value(cell) ≤ rhs
            row = np.zeros(nf); off_sum = 0.0
            for k, co in cd.items():
                vk, ok = vec(k); row += co * vk; off_sum += co * ok
            crows.append((row, rhs - off_sum))
        tgt = np.array([float(rand_val(k)) for k in free])
        # HARD LP first (vars [f, aux]); objective = L1-close to a random target.
        A_ub = [np.concatenate([row, np.zeros(nf)]) for row, _ in crows]
        b_ub = [rhs for _, rhs in crows]
        for i in range(nf):
            rp = np.zeros(2 * nf); rp[i] = 1; rp[nf + i] = -1; A_ub.append(rp); b_ub.append(tgt[i])
            rn = np.zeros(2 * nf); rn[i] = -1; rn[nf + i] = -1; A_ub.append(rn); b_ub.append(-tgt[i])
        obj = np.concatenate([np.zeros(nf), np.ones(nf)])
        bounds = [(0, 1e9)] * nf + [(0, None)] * nf
        try:
            sol = linprog(obj, A_ub=np.array(A_ub), b_ub=np.array(b_ub), bounds=bounds, method="highs")
        except Exception:
            sol = None
        if sol is None or not sol.success:
            # SOFT fallback for infeasible components: a slack ≥0 per constraint (heavily penalised) so
            # the LP always solves and MINIMISES total violation instead of falling back to random.
            ncr = len(crows); W = 2 * nf + ncr
            A2 = []; b2 = []
            for j, (row, rhs) in enumerate(crows):
                full = np.zeros(W); full[:nf] = row; full[2 * nf + j] = -1.0
                A2.append(full); b2.append(rhs)
            for i in range(nf):
                rp = np.zeros(W); rp[i] = 1; rp[nf + i] = -1; A2.append(rp); b2.append(tgt[i])
                rn = np.zeros(W); rn[i] = -1; rn[nf + i] = -1; A2.append(rn); b2.append(-tgt[i])
            obj2 = np.zeros(W); obj2[nf:2 * nf] = 1.0; obj2[2 * nf:] = 1e7
            bounds2 = [(0, 1e9)] * nf + [(0, None)] * nf + [(0, None)] * ncr
            try:
                sol = linprog(obj2, A_ub=np.array(A2), b_ub=np.array(b2), bounds=bounds2, method="highs")
            except Exception:
                sol = None
        if sol is not None and sol.success:
            # The LP guarantees every cell ≥ 0 at its (continuous) optimum. Snapping free vars to
            # multiples of 1000 is purely cosmetic but PERTURBS that point, so a derived pivot
            # (= nb − Σ coef·free) can go negative — esp. with fractional coefficients (b0471's 0.9
            # → values like −200/−300). The user's hard rule is NO negative datapoint unless a rule
            # mandates < 0. So: try the clean snapped values, accept them ONLY if every cell stays
            # ≥ 0; otherwise fall back to the LP's continuous ≥0 solution (integers), clamping any
            # residual (an infeasible/soft-slack component) to 0. ≥0 always wins over clean/exact.
            snapped = {k: F(int(round(sol.x[fidx[k]] / 1000)) * 1000) for k in free}
            cont = {k: F(max(0, int(round(sol.x[fidx[k]])))) for k in free}
            chosen = None
            for trial in (snapped, cont):
                for k in free:
                    val[k] = trial[k]
                derived = {pv: nb - sum(c * val.get(k, F(0)) for k, c in nc.items() if k != pv)
                           for pv, (nc, nb) in piv.items()}
                if all(v >= 0 for v in trial.values()) and all(v >= 0 for v in derived.values()):
                    chosen = (trial, derived)
                    break
            if chosen is None:                          # genuinely infeasible (soft-slack) → clamp ≥0
                for k in free:
                    val[k] = cont[k]
                for pv, (nc, nb) in piv.items():
                    dv = nb - sum(c * val.get(k, F(0)) for k, c in nc.items() if k != pv)
                    val[pv] = dv if dv >= 0 else F(0)
            else:
                trial, derived = chosen
                for k in free:
                    val[k] = trial[k]
                for pv, dv in derived.items():
                    val[pv] = dv
        else:
            for k in free:
                val[k] = rand_val(k)
            # no-LP fallback (scipy down / component > cap): derive pivots, but never emit a negative
            # datapoint — floor derived cells at 0 (best-effort; this degraded path skips inequalities).
            for pv, (nc, nb) in piv.items():
                dv = nb - sum(c * val.get(k, F(0)) for k, c in nc.items() if k != pv)
                val[pv] = dv if dv >= 0 else F(0)

    out: dict = {}
    for k, v in val.items():                         # k = (concept-qname, sorted-dims-tuple)
        if not is_const(k):
            out[k] = {"concept": k[0], "dims": dict(k[1]), "value": float(v)}
    return out


def _toposort(targets: set, deps: dict) -> list:
    order, seen, stack = [], set(), set()

    def visit(n):
        if n in seen:
            return
        stack.add(n)
        for m in deps.get(n, ()):
            if m in targets and m not in stack:
                visit(m)
        stack.discard(n)
        seen.add(n)
        order.append(n)
    for n in list(targets):
        visit(n)
    return order


def solve_cells(asts: list, resolver: "CellResolver", rng, datatype_of=None, rounds: int = 3,
                null_keys: set | None = None, fixed_values: dict | None = None) -> dict:
    """Compute a value for every referenced datapoint so the equality rules hold.

    Picks one derived datapoint per equality (greedy: a free cell, preferring the lone 'total'),
    randomises the leaves, then computes derived values in dependency order. `null_keys` are cells
    an isNull rule forces empty — they are fixed at 0 and never chosen as a derived target, so
    totals computed from them stay consistent once those cells are dropped from the instance.
    `fixed_values` {key->value} are cells pinned by an inequality/constraint pre-pass (e.g. c0102≤0):
    held at their value and never derived, so additive totals (e.g. c0104=c0090+c0101+c0102) stay
    consistent with them. Returns {dp_key -> {concept, dims, value}}."""
    null_keys = null_keys or set()
    fixed_values = fixed_values or {}
    plans = [p for p in (plan_equality(a, resolver) for a in asts) if p]
    # A cell that is the 'total' (single-term side) of ANY equation has its own defining equation;
    # such cells must only ever be derived from THAT equation, never as a fallback for another — else
    # we overwrite a subtotal and break its own rule (the 2D row+column over-determined case).
    total_cells: set = set()
    for p in plans:
        total_cells.update(p["preferred"])
    derived: dict = {}           # target_key -> {coef_target, contribs:[(key,coef)]}
    all_dp: dict = {}
    deps: dict = {}
    for p in plans:
        all_dp.update(p["dp"])
        # Prefer this equation's lone 'total' cell. If it's already derived elsewhere, fall back ONLY
        # to a genuine LEAF (a cell that is no equation's total) — never to another subtotal. A
        # redundant equation left unassigned this way still holds by consistency (row- and column-
        # sums both trace to the same inner leaves). Never derive a null cell (must stay 0).
        cands = [k for k in p["preferred"] if k not in derived and k not in null_keys
                 and k not in fixed_values] + \
                [k for k in p["agg"] if k not in derived and k not in null_keys
                 and k not in fixed_values and k not in total_cells]
        chosen = next(iter(cands), None)
        if chosen is None:
            continue             # every cell already derived/null/total elsewhere -> rely on consistency
        coef_t = p["agg"][chosen]
        contribs = [(k, c) for k, c in p["agg"].items() if k != chosen]
        derived[chosen] = {"coef_target": coef_t, "contribs": contribs}
        deps[chosen] = {k for k, _ in contribs}

    def rand_leaf(key):
        concept = all_dp[key]["concept"].split(":")[-1]
        dt = (datatype_of(concept) if datatype_of else None) or "MONETARY"
        if dt == "MONETARY":
            return float((rng.randint(1, 9999)) * 1000)
        if dt == "INTEGER":
            return float(rng.randint(0, 100000))
        if dt in ("PERCENTAGE",):
            return round(rng.uniform(0, 1), 4)
        return round(rng.uniform(0, 9999), 2)

    values: dict = {}
    for key in all_dp:
        if key in null_keys:
            values[key] = 0.0                     # isNull cell — fixed empty (contributes 0)
        elif key in fixed_values:
            values[key] = fixed_values[key]       # inequality/constraint-pinned cell (held, not derived)
        elif key not in derived:
            values[key] = rand_leaf(key)
    order = _toposort(set(derived), deps)
    for _ in range(rounds):                      # a few passes so chained totals settle
        for t in order:
            p = derived[t]
            s = 0.0
            for k, c in p["contribs"]:
                s += c * values.get(k, 0.0)
            values[t] = -s / p["coef_target"]
    return {k: {"concept": all_dp[k]["concept"], "dims": all_dp[k]["dims"],
                "table": all_dp[k].get("table"), "value": values[k]}
            for k in all_dp}


# ----------------------------------------------------------------- verify main
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.path.insert(0, os.path.dirname(__file__))

    ext = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\177069\ClaudeLearning\boebanking400"
    xlsx = (r"C:\Users\177069\ClaudeLearning\boebankingtaxonomyvalidationsv400"
            r"\Bank of England Banking Taxonomy Validations Banking reporting v4.0.0.xlsx")
    rules = load_workbook_rules(xlsx, "banking_reporting")
    print(f"loaded {len(rules)} workbook rules")
    parsed = [r for r in rules if parse_expression(r["expression"])]
    print(f"parsed (additive equality/inequality shape): {len(parsed)} / {len(rules)}")

    res = CellResolver(ext)
    for code in ("v7380_m", "v7381_m", "v7382_m"):
        r = next((x for x in rules if x["code"] == code), None)
        if not r:
            continue
        ast = parse_expression(r["expression"])
        print(f"\n{code}: op={ast['op']}  lhs_terms={len(ast['lhs'])} rhs_terms={len(ast['rhs'])}")
        for side in ("lhs", "rhs"):
            for t in ast[side]:
                dps = res.resolve(t["cell"])
                tag = f"{t['cell']['table']} r{t['cell']['r']} c{t['cell']['c']} z{t['cell']['z']}"
                print(f"  [{side}] coef={t['coef']} sum={t['cell']['sum']} {tag} -> {len(dps)} datapoint(s)"
                      + (f"  e.g. {dps[0]['concept']} {dps[0]['dims']}" if dps else "  (UNRESOLVED)"))
